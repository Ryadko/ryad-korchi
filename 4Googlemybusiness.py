import requests
import json
import time
import os
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

# Charger les variables d'environnement à partir d'un fichier .env
load_dotenv()

# Récupérer la clé API depuis les variables d'environnement
API_KEY = os.getenv("GOOGLE_PLACES_API_KEY")

# Si la clé API n'est pas dans le fichier .env, utiliser celle fournie en argument
if not API_KEY:
    API_KEY = "VOTRE_CLE_API"  # Remplacez par votre clé API

def rechercher_et_obtenir_details(nom_entreprise, ville):
    """
    Effectue une recherche et obtient les détails d'une entreprise
    """
    # Construction de la requête
    query = f"{nom_entreprise} {ville}"
    
    # Recherche textuelle pour obtenir le place_id
    url_search = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params_search = {
        "query": query,
        "key": API_KEY
    }
    
    response_search = requests.get(url_search, params=params_search)
    data_search = response_search.json()
    
    # Si la recherche a fonctionné et qu'on a des résultats
    if data_search.get('status') == 'OK' and data_search.get('results'):
        place_id = data_search['results'][0]['place_id']
        
        # Obtenir les détails avec le place_id
        url_details = "https://maps.googleapis.com/maps/api/place/details/json"
        params_details = {
            "place_id": place_id,
            "fields": "name,formatted_address,formatted_phone_number,website,url",
            "key": API_KEY
        }
        
        response_details = requests.get(url_details, params=params_details)
        return response_details.json()
    
    return data_search  # Retourner les données brutes même si pas de résultats

def creer_fichier_exemple(fichier_entree):
    """
    Crée un fichier Excel d'exemple avec quelques entreprises
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Entreprises"
    
    # Ajouter l'en-tête
    ws.append(["Nom", "Ville"])
    
    # Ajouter quelques exemples
    ws.append(["Boulangerie Martin", "Paris"])
    ws.append(["Restaurant Le Gourmet", "Lyon"])
    ws.append(["Garage Auto Express", "Marseille"])
    
    # Sauvegarder le fichier
    wb.save(fichier_entree)
    print(f"Fichier exemple '{fichier_entree}' créé.")

def traiter_liste_entreprises(fichier_entree, fichier_sortie):
    """
    Traite une liste d'entreprises à partir d'un fichier Excel et écrit les résultats dans un autre fichier Excel
    """
    # Créer un exemple si le fichier n'existe pas
    if not os.path.exists(fichier_entree):
        creer_fichier_exemple(fichier_entree)
    
    # Lire le fichier d'entrée Excel
    wb_entree = load_workbook(filename=fichier_entree)
    ws_entree = wb_entree.active
    
    # Créer un nouveau classeur pour les résultats
    wb_sortie = Workbook()
    ws_sortie = wb_sortie.active
    ws_sortie.title = "Résultats"
    
    # Écrire l'en-tête dans le fichier de sortie
    ws_sortie.append(['Nom', 'Ville', 'Adresse', 'Téléphone', 'Site Web', 'URL Google Maps', 'Données Brutes'])
    
    # Ignorer la première ligne (en-tête)
    rows = list(ws_entree.rows)
    entreprises = [(row[0].value, row[1].value) for row in rows[1:] if row[0].value and row[1].value]
    
    # Traiter chaque entreprise
    for nom, ville in entreprises:
        print(f"Recherche de {nom} à {ville}...")
        
        # Rechercher et obtenir les détails
        resultat = rechercher_et_obtenir_details(nom, ville)
        
        # Extraire les informations si disponibles
        adresse = ""
        telephone = ""
        site_web = ""
        maps_url = ""
        
        # Récupérer les données brutes
        donnees_brutes = json.dumps(resultat, ensure_ascii=False)
        
        # Si nous avons un résultat valide
        if resultat.get('status') == 'OK' and resultat.get('result'):
            info = resultat['result']
            adresse = info.get('formatted_address', '')
            telephone = info.get('formatted_phone_number', '')
            site_web = info.get('website', '')
            maps_url = info.get('url', '')
        
        # Écrire les informations dans le fichier Excel
        ws_sortie.append([nom, ville, adresse, telephone, site_web, maps_url, donnees_brutes])
        
        # Attendre un peu pour éviter de dépasser les limites de quota de l'API
        time.sleep(0.2)
    
    # Ajuster la largeur des colonnes
    for col in ws_sortie.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        adjusted_width = min(max_length + 2, 100)  # Limiter la largeur à 100 pour éviter des colonnes trop larges
        ws_sortie.column_dimensions[column].width = adjusted_width
    
    # Sauvegarder le fichier de résultats
    wb_sortie.save(fichier_sortie)
            
if __name__ == "__main__":
    # Fichiers d'entrée et de sortie
    FICHIER_ENTREE = "entreprises.xlsx"
    FICHIER_SORTIE = "resultats_entreprises.xlsx"
    
    # Traiter la liste d'entreprises
    traiter_liste_entreprises(FICHIER_ENTREE, FICHIER_SORTIE)
    print(f"Traitement terminé. Résultats enregistrés dans '{FICHIER_SORTIE}'")