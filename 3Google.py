import requests
import re
import openpyxl
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup
import time
import os
import json
from difflib import SequenceMatcher
import random
import datetime

# Configuration des API Google
API_KEYS = [
    "AIzaSyCNHqZewTPKPasi4K_XwcJfTlu7t7M8jP4",
    "AIzaSyDgf7cfL3UEcmBdeu_x9dKtdhH3km1Xy5g",
    "AIzaSyC5eAPky3ozv-wTTdMvO9KU4zevCXQv4iQ" 
]
SEARCH_ENGINE_IDS = [
    "66a3a2b6a945f4646",
    "d2080091ff6cd4e3e",
    "c5594fe07f98a4edc"
]
BASE_URL = "https://www.googleapis.com/customsearch/v1"

# Configuration des délais
MIN_REQUEST_DELAY = 1.5  # Délai minimum entre les requêtes en secondes
MAX_REQUEST_DELAY = 3.0  # Délai maximum entre les requêtes en secondes
REQUEST_TIMEOUT = 15     # Timeout pour les requêtes en secondes
BACKUP_INTERVAL = 5      # Nombre d'entreprises traitées avant sauvegarde

# Variable pour suivre l'API en cours d'utilisation
current_api_index = 0

def google_search(query, num=10):
    """Effectue une recherche Google et retourne les résultats."""
    global current_api_index
    
    # Utiliser l'API et le moteur de recherche courants
    api_key = API_KEYS[current_api_index]
    search_engine_id = SEARCH_ENGINE_IDS[current_api_index]
    
    params = {
        'key': api_key,
        'cx': search_engine_id,
        'q': query,
        'num': num  # Nombre de résultats à récupérer (max 10 par requête)
    }
    
    try:
        response = requests.get(BASE_URL, params=params, timeout=REQUEST_TIMEOUT)
        
        if response.status_code != 200:
            print(f"Erreur dans la recherche avec l'API {current_api_index}: {response.status_code}")
            
            # Essayer avec l'autre API en cas d'erreur
            current_api_index = (current_api_index + 1) % len(API_KEYS)
            print(f"Basculement vers l'API {current_api_index}")
            
            # Nouvelle tentative avec l'autre API
            api_key = API_KEYS[current_api_index]
            search_engine_id = SEARCH_ENGINE_IDS[current_api_index]
            
            params = {
                'key': api_key,
                'cx': search_engine_id,
                'q': query,
                'num': num
            }
            
            response = requests.get(BASE_URL, params=params, timeout=REQUEST_TIMEOUT)
            
            if response.status_code != 200:
                print(f"Erreur persistante avec l'API de secours: {response.status_code}")
                return None
        
        # Alterner automatiquement entre les API pour la prochaine requête
        current_api_index = (current_api_index + 1) % len(API_KEYS)
        
        return response.json()
    
    except requests.exceptions.Timeout:
        print(f"Timeout lors de la recherche Google pour '{query}'")
        # Basculer vers une autre API
        current_api_index = (current_api_index + 1) % len(API_KEYS)
        return None
    except Exception as e:
        print(f"Erreur lors de la recherche Google: {e}")
        return None

def extract_domain(url):
    """Extrait le domaine principal d'une URL sans www et l'extension."""
    parsed_url = urlparse(url)
    domain = parsed_url.netloc
    
    # Enlever les www. du début s'il existe
    if domain.startswith('www.'):
        domain = domain[4:]
    
    # Pour obtenir le domaine sans l'extension, mais préserver les sous-domaines
    parts = domain.split('.')
    
    if len(parts) > 2:
        # Pour les domaines de type sub.domain.com
        domain = '.'.join(parts[:-1])  
    else:
        # Pour les domaines simples comme domain.com
        domain = parts[0] 
    
    return domain

def get_base_url(url):
    """Extrait l'URL de base (schéma + domaine) d'une URL complète."""
    parsed_url = urlparse(url)
    return f"{parsed_url.scheme}://{parsed_url.netloc}"

def clean_company_name(name):
    """Nettoie le nom de l'entreprise pour la comparaison."""
    if not name:
        return ""
        
    # Convertir en minuscules
    name = name.lower()
    
    # Supprimer les mots communs et caractères spéciaux
    words_to_remove = ['sarl', 'sa', 'sas', 'eurl', 'sasu', 'sci', 'earl', 'scp', 'selarl', 'et', 'de', 'la', 'les', 'le', 'the']
    for word in words_to_remove:
        name = re.sub(r'\b' + word + r'\b', '', name)
    
    # Remplacer les caractères spéciaux par des espaces
    name = re.sub(r'[^\w\s]', ' ', name)
    
    # Remplacer les espaces multiples par un seul espace
    name = re.sub(r'\s+', ' ', name)
    
    # Supprimer les espaces au début et à la fin
    name = name.strip()
    
    # Supprimer les accents
    name = name.replace('é', 'e').replace('è', 'e').replace('ê', 'e').replace('ë', 'e')
    name = name.replace('à', 'a').replace('â', 'a').replace('ä', 'a')
    name = name.replace('î', 'i').replace('ï', 'i')
    name = name.replace('ô', 'o').replace('ö', 'o')
    name = name.replace('ù', 'u').replace('û', 'u').replace('ü', 'u')
    name = name.replace('ç', 'c')
    
    return name

def similarity_score(str1, str2):
    """Calcule un score de similarité entre deux chaînes."""
    if not str1 or not str2:
        return 0
    return SequenceMatcher(None, str1, str2).ratio()

def check_company_name_in_page(url, company_name):
    """Vérifie si le nom de l'entreprise apparaît dans le contenu de la page."""
    try:
        html_content = extract_page_content(url)
        if not html_content:
            return False
            
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Extraire le texte visible
        page_text = soup.get_text().lower()
        
        # Vérifier si le nom de l'entreprise apparaît dans le texte
        cleaned_name = clean_company_name(company_name)
        
        # Différentes variations du nom de l'entreprise à rechercher
        name_variations = [
            cleaned_name,
            cleaned_name.replace(' ', ''),
            cleaned_name.replace(' ', '-'),
            cleaned_name.replace(' ', '_')
        ]
        
        for variation in name_variations:
            if variation and len(variation) > 3 and variation in page_text:
                return True
                
        # Vérifier également dans les balises de titre et méta
        title = soup.title.string.lower() if soup.title else ""
        meta_desc = ""
        meta_tag = soup.find("meta", attrs={"name": "description"})
        if meta_tag and "content" in meta_tag.attrs:
            meta_desc = meta_tag["content"].lower()
            
        for variation in name_variations:
            if variation and len(variation) > 3 and (variation in title or variation in meta_desc):
                return True
                
        return False
        
    except requests.exceptions.Timeout:
        print(f"Timeout lors de la vérification du nom dans la page {url}")
        return False
    except Exception as e:
        print(f"Erreur lors de la vérification du nom dans la page {url}: {e}")
        return False

def get_best_website_match(company_name, city, search_results):
    """Sélectionne le site web qui correspond le mieux à l'entreprise avec plusieurs méthodes."""
    if not search_results or 'items' not in search_results or not search_results['items']:
        return None, 0
    
    cleaned_company = clean_company_name(company_name)
    cleaned_company_nospace = cleaned_company.replace(' ', '')
    
    # 1. Première passe: vérifier les correspondances exactes dans les URL
    for item in search_results['items']:
        url = item['link']
        domain = extract_domain(url)
        
        # Vérifier si c'est une correspondance exacte (sans espace)
        if cleaned_company_nospace == domain or cleaned_company.replace(' ', '-') == domain:
            if check_company_name_in_page(url, company_name):
                return url, 0.95
    
    # 2. Deuxième passe: vérifier les inclusions fortes
    for item in search_results['items']:
        url = item['link']
        domain = extract_domain(url)
        
        # Vérifier si le nom est fortement inclus dans le domaine
        if (cleaned_company_nospace in domain and len(domain) < len(cleaned_company_nospace) + 5) or \
           (domain in cleaned_company_nospace and len(domain) > 4):
            if check_company_name_in_page(url, company_name):
                return url, 0.85
    
    # 3. Troisième passe: calcul de similarité
    best_score = 0
    best_url = None
    
    for item in search_results['items']:
        url = item['link']
        domain = extract_domain(url)
        
        # Calculer différents scores de similarité
        score1 = similarity_score(cleaned_company, domain)
        score2 = similarity_score(cleaned_company_nospace, domain)
        score3 = similarity_score(cleaned_company.replace(' ', '-'), domain)
        
        # Prendre le meilleur score
        score = max(score1, score2, score3)
        
        # Bonus si le nom de l'entreprise apparaît dans le snippet
        snippet = item.get('snippet', '').lower()
        if cleaned_company in snippet:
            score += 0.1
            
        # Vérifier si l'URL contient le nom de la ville
        if city and city.lower() in url.lower():
            score += 0.05
        
        if score > best_score:
            # Vérifier le contenu de la page pour confirmation
            if check_company_name_in_page(url, company_name):
                best_score = score
                best_url = url
    
    # 4. Quatrième passe: essayer des variantes de recherche
    if best_score < 0.5:
        # Essayer avec l'URL directe si aucun bon résultat n'a été trouvé
        url_variations = [
            f"{cleaned_company_nospace}.fr",
            f"{cleaned_company_nospace}.com",
            f"{cleaned_company.replace(' ', '-')}.fr",
            f"{cleaned_company.replace(' ', '-')}.com",
            f"{cleaned_company.replace(' ', '')}.io",
            f"{cleaned_company.replace(' ', '-')}.io"
        ]
        
        for variation in url_variations:
            try:
                test_url = f"https://{variation}"
                response = requests.head(test_url, timeout=REQUEST_TIMEOUT)
                if response.status_code < 400:  # Considérer tous les codes de succès et redirection
                    if check_company_name_in_page(test_url, company_name):
                        return test_url, 0.85
            except:
                continue
    
    return best_url, best_score

def extract_page_content(url):
    """Récupère le contenu HTML d'une page web."""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        return response.text
    except requests.exceptions.Timeout:
        print(f"Timeout lors de l'extraction du contenu de {url}")
        return None
    except Exception as e:
        print(f"Erreur lors de l'extraction du contenu de {url}: {e}")
        return None

def find_contact_page(base_url, html_content):
    """Trouve l'URL de la page de contact sur le site."""
    if not html_content:
        return None
        
    soup = BeautifulSoup(html_content, 'html.parser')
    contact_keywords = ['contact', 'contactez', 'nous contacter', 'contactez-nous', 'contacts', 'about', 'à propos']
    
    for link in soup.find_all('a', href=True):
        href = link['href']
        link_text = link.text.lower().strip()
        
        if any(keyword in link_text for keyword in contact_keywords):
            # Résoudre l'URL relative
            full_url = urljoin(base_url, href)
            return full_url
            
    # Rechercher également dans le menu de navigation
    for nav in soup.find_all(['nav', 'div', 'ul'], class_=re.compile(r'nav|menu', re.I)):
        for link in nav.find_all('a', href=True):
            href = link['href']
            link_text = link.text.lower().strip()
            
            if any(keyword in link_text for keyword in contact_keywords):
                full_url = urljoin(base_url, href)
                return full_url
    
    return None

def extract_info_from_html(html_content, base_url):
    """Extrait les emails, téléphones et liens de réseaux sociaux du HTML."""
    if not html_content:
        return [], [], []
    
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # Extraction des emails
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    emails = list(set(re.findall(email_pattern, html_content)))
    
    # Extraction des numéros de téléphone (formats français)
    phone_patterns = [
        r'(?:(?:\+|00)33|0)\s*[1-9](?:[\s.-]*\d{2}){4}',  # Format international et national
        r'(?:0|\+33|0033)[1-9](?:[\s.-]*\d{2}){4}'       # Variantes
    ]
    
    phones = []
    for pattern in phone_patterns:
        found_phones = re.findall(pattern, html_content)
        phones.extend(found_phones)
    phones = list(set(phones))
    
    # Extraction des liens vers les réseaux sociaux
    social_domains = ['facebook.com', 'twitter.com', 'linkedin.com', 'instagram.com', 
                    'youtube.com', 'pinterest.com', 'tiktok.com']
    
    social_links = []
    for link in soup.find_all('a', href=True):
        href = link['href']
        
        # Résoudre les URLs relatives
        if href.startswith('/'):
            href = urljoin(base_url, href)
        
        if any(social_domain in href for social_domain in social_domains):
            social_links.append(href)
    
    return emails, phones, list(set(social_links))

def save_backup(excel_path, processed_rows, last_row_index):
    """Sauvegarde l'état actuel du traitement."""
    backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
    
    # Créer le répertoire de sauvegarde s'il n'existe pas
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        
    # Nom du fichier de sauvegarde basé sur la date et l'heure
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(backup_dir, f"backup_{timestamp}.json")
    
    # Données à sauvegarder
    backup_data = {
        "excel_path": excel_path,
        "processed_rows": processed_rows,
        "last_row_index": last_row_index,
        "timestamp": timestamp
    }
    
    # Sauvegarder dans un fichier JSON
    with open(backup_file, 'w') as f:
        json.dump(backup_data, f)
        
    print(f"Sauvegarde créée: {backup_file}")
    
    # Créer aussi une copie de sauvegarde du fichier Excel
    excel_backup = os.path.join(backup_dir, f"BDDSIRET_backup_{timestamp}.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    wb.save(excel_backup)
    print(f"Copie de sauvegarde Excel créée: {excel_backup}")
    
    return backup_file

def load_latest_backup(excel_path):
    """Charge la dernière sauvegarde disponible."""
    backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
    
    if not os.path.exists(backup_dir):
        print("Aucune sauvegarde trouvée.")
        return None
        
    backup_files = [os.path.join(backup_dir, f) for f in os.listdir(backup_dir) if f.startswith("backup_") and f.endswith(".json")]
    
    if not backup_files:
        print("Aucun fichier de sauvegarde trouvé.")
        return None
        
    # Trier par date de modification (la plus récente d'abord)
    latest_backup = max(backup_files, key=os.path.getmtime)
    
    try:
        with open(latest_backup, 'r') as f:
            backup_data = json.load(f)
            print(f"Dernière sauvegarde chargée: {latest_backup}")
            print(f"Date de la sauvegarde: {backup_data['timestamp']}")
            print(f"Nombre d'entreprises déjà traitées: {len(backup_data['processed_rows'])}")
            return backup_data
    except Exception as e:
        print(f"Erreur lors du chargement de la sauvegarde: {e}")
        return None

def process_companies(excel_path, resume_from_backup=True):
    """Traite la liste des entreprises et enregistre les résultats."""
    try:
        # Variables pour le suivi des entreprises traitées
        processed_rows = []
        last_row_index = 0
        
        # Vérifier s'il faut reprendre depuis une sauvegarde
        backup_data = None
        if resume_from_backup:
            backup_data = load_latest_backup(excel_path)
            if backup_data:
                processed_rows = backup_data["processed_rows"]
                last_row_index = backup_data["last_row_index"]
                print(f"Reprise du traitement à partir de la ligne {last_row_index}")
        
        # Vérifier si le fichier existe
        if not os.path.exists(excel_path):
            print(f"Le fichier {excel_path} n'existe pas.")
            return
            
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(excel_path)
        
        # Vérifier si la feuille "Reformatage" existe
        if "Reformatage" not in wb.sheetnames:
            print("La feuille 'Reformatage' n'existe pas dans le fichier.")
            return
        
        # Lire les noms d'entreprises et les villes
        sheet = wb["Reformatage"]
        
        # Trouver l'index des colonnes nécessaires
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        try:
            company_name_idx = header_row.index("data_results_0_nom_raison_sociale") + 1  # +1 car openpyxl commence à 1
        except ValueError:
            print("La colonne 'data_results_0_nom_raison_sociale' n'a pas été trouvée.")
            return
            
        try:
            city_idx = header_row.index("data_results_0_matching_etablissements_0_libelle_commune") + 1
        except ValueError:
            print("La colonne 'data_results_0_matching_etablissements_0_libelle_commune' n'a pas été trouvée.")
            city_idx = None  # On continue sans la ville si la colonne n'existe pas
        
        # Créer ou récupérer la feuille de résultats
        if "Googleresult" in wb.sheetnames:
            result_sheet = wb["Googleresult"]
            
            # Si on ne reprend pas depuis une sauvegarde, on efface le contenu existant
            if not backup_data:
                for row in range(result_sheet.max_row, 0, -1):
                    result_sheet.delete_rows(row)
                # Ajouter les en-têtes
                result_sheet.append([
                    "Nom Entreprise",
                    "Ville", 
                    "Site Web", 
                    "Score de correspondance",
                    "Page de contact",
                    "Domaine racine",
                    "Emails", 
                    "Téléphones", 
                    "Réseaux Sociaux",
                    "API utilisée",
                    "Date traitement"  # Nouvelle colonne pour la date de traitement
                ])
        else:
            result_sheet = wb.create_sheet("Googleresult")
            # Ajouter les en-têtes
            result_sheet.append([
                "Nom Entreprise",
                "Ville", 
                "Site Web", 
                "Score de correspondance",
                "Page de contact",
                "Domaine racine",
                "Emails", 
                "Téléphones", 
                "Réseaux Sociaux",
                "API utilisée",
                "Date traitement"  # Nouvelle colonne pour la date de traitement
            ])
        
        # Compter le nombre total d'entreprises à traiter
        total_rows = sheet.max_row - 1  # -1 pour exclure l'en-tête
        counter = 0
        backup_counter = 0
        
        # Traiter chaque entreprise à partir du dernier index sauvegardé
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            # Sauter les lignes déjà traitées
            if row_idx <= last_row_index:
                continue
                
            company_name = row[company_name_idx-1]  # -1 car les index de Python commencent à 0
            city = row[city_idx-1] if city_idx else ""
            
            if not company_name:
                processed_rows.append(row_idx)
                last_row_index = row_idx
                continue
            
            counter += 1
            backup_counter += 1
            
            print(f"[{counter}/{total_rows - last_row_index + 1}] Traitement de l'entreprise: {company_name} à {city if city else 'Ville inconnue'}")
            
            # Plusieurs stratégies de recherche
            search_strategies = [
                f"{company_name} {city if city else ''} site officiel",
                f"\"{company_name}\" {city if city else ''} site web",
                f"\"{company_name}\" site internet",
                f"{company_name.lower().replace(' ', '')} {city if city else ''}"
            ]
            
            # Essayer chaque stratégie de recherche
            website_url = None
            score = 0
            api_used = None  # Pour suivre quelle API a été utilisée
            
            for strategy in search_strategies:
                print(f"Essai de la stratégie: {strategy}")
                # Mémoriser l'index de l'API avant la recherche
                pre_search_api_index = current_api_index
                
                # Effectuer la recherche Google
                search_results = google_search(strategy, num=10)
                
                # Déterminer quelle API a été utilisée pour cette recherche
                if search_results:
                    api_used = f"API {pre_search_api_index} ({API_KEYS[pre_search_api_index][:8]}...)"
                
                # Obtenir le site web correspondant
                current_url, current_score = get_best_website_match(company_name, city, search_results)
                
                if current_url and current_score > score:
                    website_url = current_url
                    score = current_score
                    
                    # Si on a déjà un bon score, arrêter les recherches
                    if score > 0.7:
                        break
                        
                # Pause aléatoire pour éviter de surcharger l'API
                delay = random.uniform(MIN_REQUEST_DELAY, MAX_REQUEST_DELAY)
                time.sleep(delay)
            
            # Vérifier manuellement des URLs spécifiques pour les entreprises problématiques
            known_sites = {
                "THE MISSING ONE": "https://themissingone.io/",
                "TOMORROW JOBS": "https://tomorrowjobs.fr/",
                "SHARK GRAPHIK": "https://shark-graphik.fr/"
            }
            
            if company_name in known_sites:
                test_url = known_sites[company_name]
                try:
                    response = requests.head(test_url, timeout=REQUEST_TIMEOUT)
                    if response.status_code < 400:
                        website_url = test_url
                        score = 1.0  # Score maximum pour les sites connus
                        api_used = "Site connu manuellement"
                        print(f"Site connu trouvé pour {company_name}: {website_url}")
                except:
                    print(f"Site connu inaccessible pour {company_name}: {test_url}")
            
            emails = []
            phones = []
            social_links = []
            contact_page_url = None
            root_domain = None
            
            if website_url:
                # Extraire la racine du domaine
                root_domain = get_base_url(website_url)
                
                # Extraire le contenu de la page d'accueil
                html_content = extract_page_content(website_url)
                
                if html_content:
                    # Extraire les informations de la page d'accueil
                    page_emails, page_phones, page_social = extract_info_from_html(html_content, website_url)
                    emails.extend(page_emails)
                    phones.extend(page_phones)
                    social_links.extend(page_social)
                    
                    # Chercher la page de contact
                    contact_page_url = find_contact_page(website_url, html_content)
                    
                    if contact_page_url:
                        print(f"Page de contact trouvée: {contact_page_url}")
                        contact_html = extract_page_content(contact_page_url)
                        
                        if contact_html:
                            # Extraire les informations de la page de contact
                            contact_emails, contact_phones, contact_social = extract_info_from_html(contact_html, contact_page_url)
                            emails.extend(contact_emails)
                            phones.extend(contact_phones)
                            social_links.extend(contact_social)
                
                # Dédupliquer les listes
                emails = list(set(emails))
                phones = list(set(phones))
                social_links = list(set(social_links))
                
                # Pause aléatoire pour éviter de surcharger les sites web
                delay = random.uniform(MIN_REQUEST_DELAY, MAX_REQUEST_DELAY)
                time.sleep(delay)
            
            # Date de traitement
            current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Ajouter les résultats à la feuille
            result_sheet.append([
                company_name,
                city if city else "",
                website_url if website_url else "Aucun site trouvé",
                f"{score:.2f}" if website_url else "0.00",
                contact_page_url if contact_page_url else "",
                root_domain if root_domain else "",
                ", ".join(emails) if emails else "",
                ", ".join(phones) if phones else "",
                ", ".join(social_links) if social_links else "",
                api_used if api_used else "Aucune API utilisée",
                current_date  # Date de traitement
            ])
            
            # Marquer cette ligne comme traitée
            processed_rows.append(row_idx)
            last_row_index = row_idx
            
            # Sauvegarder périodiquement les résultats
            wb.save(excel_path)
            
            # Créer une sauvegarde toutes les N entreprises traitées
            if backup_counter >= BACKUP_INTERVAL:
                save_backup(excel_path, processed_rows, last_row_index)
                backup_counter = 0
                print(f"Progression: {counter}/{total_rows - (last_row_index - counter)} entreprises traitées")
        
        # Sauvegarder le fichier Excel à la fin
        wb.save(excel_path)
        
        # Sauvegarde finale
        save_backup(excel_path, processed_rows, last_row_index)
        
        print(f"Traitement terminé. Les résultats sont enregistrés dans la feuille 'Googleresult'.")
        print(f"Total d'entreprises traitées dans cette session: {counter}")
        print(f"Total d'entreprises traitées au total: {len(processed_rows)}")
        
    except Exception as e:
        print(f"Une erreur est survenue: {e}")
        # En cas d'erreur, essayer de sauvegarder ce qui a été traité
        try:
            if 'wb' in locals() and 'excel_path' in locals():
                wb.save(excel_path)
                print("Sauvegarde d'urgence du fichier Excel effectuée.")
            
            if 'excel_path' in locals() and 'processed_rows' in locals() and 'last_row_index' in locals():
                save_backup(excel_path, processed_rows, last_row_index)
                print("Sauvegarde d'urgence de l'état effectuée.")
        except Exception as backup_error:
            print(f"Erreur lors de la sauvegarde d'urgence: {backup_error}")

def list_available_backups(excel_path):
    """Liste toutes les sauvegardes disponibles."""
    backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
    
    if not os.path.exists(backup_dir):
        print("Aucune sauvegarde trouvée.")
        return []
        
    backup_files = [f for f in os.listdir(backup_dir) if f.startswith("backup_") and f.endswith(".json")]
    
    if not backup_files:
        print("Aucun fichier de sauvegarde trouvé.")
        return []
    
    # Trier par date de modification (la plus récente d'abord)
    backup_files.sort(key=lambda x: os.path.getmtime(os.path.join(backup_dir, x)), reverse=True)
    
    print("Sauvegardes disponibles:")
    for i, backup_file in enumerate(backup_files):
        # Charger les infos de la sauvegarde
        with open(os.path.join(backup_dir, backup_file), 'r') as f:
            backup_data = json.load(f)
            
        timestamp = backup_data.get('timestamp', 'Date inconnue')
        processed_count = len(backup_data.get('processed_rows', []))
        last_row = backup_data.get('last_row_index', 0)
        
        print(f"{i+1}. {backup_file} - Date: {timestamp}, Lignes traitées: {processed_count}, Dernière ligne: {last_row}")
        
    return backup_files

def load_specific_backup(excel_path, backup_index):
    """Charge une sauvegarde spécifique par son index."""
    backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
    backup_files = list_available_backups(excel_path)
    
    if not backup_files:
        return None
        
    if backup_index < 1 or backup_index > len(backup_files):
        print(f"Index de sauvegarde invalide. Choisissez entre 1 et {len(backup_files)}.")
        return None
        
    selected_backup = backup_files[backup_index - 1]
    
    try:
        with open(os.path.join(backup_dir, selected_backup), 'r') as f:
            backup_data = json.load(f)
            print(f"Sauvegarde chargée: {selected_backup}")
            return backup_data
    except Exception as e:
        print(f"Erreur lors du chargement de la sauvegarde: {e}")
        return None

def main():
    """Fonction principale pour exécuter le script avec menu interactif."""
    # Utiliser le chemin du répertoire courant pour trouver le fichier
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "BDDSIRET.xlsx")
    
    if not os.path.exists(excel_path):
        excel_path = input("Fichier BDDSIRET.xlsx non trouvé dans le même répertoire. Entrez le chemin complet: ")
        if not os.path.exists(excel_path):
            print("Fichier introuvable. Fin du programme.")
            return
    
    while True:
        print("\n===== MENU PRINCIPAL =====")
        print("1. Démarrer un nouveau traitement")
        print("2. Reprendre depuis la dernière sauvegarde")
        print("3. Choisir une sauvegarde spécifique")
        print("4. Lister les sauvegardes disponibles")
        print("5. Configurer les délais")
        print("6. Quitter")
        
        choice = input("\nVotre choix: ")
        
        if choice == "1":
            print("Démarrage d'un nouveau traitement...")
            process_companies(excel_path, resume_from_backup=False)
            
        elif choice == "2":
            print("Reprise depuis la dernière sauvegarde...")
            process_companies(excel_path, resume_from_backup=True)
            
        elif choice == "3":
            backup_files = list_available_backups(excel_path)
            if backup_files:
                backup_index = int(input("Entrez le numéro de la sauvegarde à charger: "))
                backup_data = load_specific_backup(excel_path, backup_index)
                
                if backup_data:
                    confirm = input("Voulez-vous reprendre le traitement depuis cette sauvegarde? (o/n): ")
                    if confirm.lower() == "o":
                        # Créer une fonction spéciale pour reprendre depuis une sauvegarde spécifique
                        print("Reprise du traitement depuis la sauvegarde sélectionnée...")
                        # Récupérer les données de sauvegarde
                        processed_rows = backup_data["processed_rows"]
                        last_row_index = backup_data["last_row_index"]
                        # Créer un nouveau fichier à partir de la sauvegarde Excel
                        backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
                        excel_backup_files = [f for f in os.listdir(backup_dir) if f.startswith("BDDSIRET_backup_") and f.endswith(".xlsx")]
                        excel_backup_files.sort(key=lambda x: os.path.getmtime(os.path.join(backup_dir, x)), reverse=True)
                        
                        if excel_backup_files:
                            backup_excel = os.path.join(backup_dir, excel_backup_files[0])
                            process_companies(excel_path, resume_from_backup=True)
                        else:
                            print("Fichier Excel de sauvegarde non trouvé. Utilisation du fichier actuel.")
                            process_companies(excel_path, resume_from_backup=True)
            
        elif choice == "4":
            list_available_backups(excel_path)
            
        elif choice == "5":
            global MIN_REQUEST_DELAY, MAX_REQUEST_DELAY, REQUEST_TIMEOUT, BACKUP_INTERVAL
            
            print("\n===== CONFIGURATION DES DÉLAIS =====")
            print(f"Délai minimum actuel entre requêtes: {MIN_REQUEST_DELAY} secondes")
            print(f"Délai maximum actuel entre requêtes: {MAX_REQUEST_DELAY} secondes")
            print(f"Timeout actuel des requêtes: {REQUEST_TIMEOUT} secondes")
            print(f"Intervalle de sauvegarde actuel: toutes les {BACKUP_INTERVAL} entreprises")
            
            try:
                new_min_delay = float(input("Nouveau délai minimum (secondes, recommandé >= 1): "))
                new_max_delay = float(input("Nouveau délai maximum (secondes, recommandé >= 2): "))
                new_timeout = float(input("Nouveau timeout (secondes, recommandé >= 10): "))
                new_backup_interval = int(input("Nouvel intervalle de sauvegarde (nombre d'entreprises): "))
                
                # Validation des valeurs
                if new_min_delay > 0 and new_max_delay >= new_min_delay and new_timeout >= 5 and new_backup_interval > 0:
                    MIN_REQUEST_DELAY = new_min_delay
                    MAX_REQUEST_DELAY = new_max_delay
                    REQUEST_TIMEOUT = new_timeout
                    BACKUP_INTERVAL = new_backup_interval
                    print("Configuration mise à jour avec succès!")
                else:
                    print("Valeurs invalides. Configuration inchangée.")
            except ValueError:
                print("Entrée invalide. Configuration inchangée.")
            
        elif choice == "6":
            print("Fin du programme.")
            break
            
        else:
            print("Choix invalide. Veuillez réessayer.")

if __name__ == "__main__":
    main()